#!/usr/bin/env python3
"""Prepare a Tecman ticket migration draft from an InvGate Excel export.

The script is intentionally conservative: by default it writes a draft JSON and
does not modify Tecman's runtime data. Use --apply to merge into data/tickets.json
after reviewing the draft.
"""

from __future__ import annotations

import argparse
import datetime as dt
import json
import re
import shutil
from collections import Counter
from pathlib import Path
from typing import Any

from openpyxl import load_workbook


CLOSED_STATES = {"cerrado", "cerrar juli", "cancelado", "rechazado"}


def clean(value: Any) -> str:
    if value is None:
        return ""
    return " ".join(str(value).replace("\xa0", " ").split()).strip()


def iso_date(value: Any) -> str:
    if not value:
        return ""
    if isinstance(value, dt.datetime):
        return value.isoformat()
    if isinstance(value, dt.date):
        return dt.datetime.combine(value, dt.time()).isoformat()
    text = clean(value)
    for fmt in ("%m/%d/%Y", "%d/%m/%Y", "%d-%m-%Y", "%d-%b-%Y"):
        try:
            return dt.datetime.strptime(text.strip(), fmt).isoformat()
        except ValueError:
            pass
    return text


def norm_sucursal(value: Any) -> tuple[str, str]:
    text = clean(value)
    if not text:
        return "", ""
    if isinstance(value, int):
        return f"Sucursal {value:03d}", f"{value:03d}"
    if isinstance(value, float) and value.is_integer():
        number = int(value)
        return f"Sucursal {number:03d}", f"{number:03d}"
    match = re.fullmatch(r"0*(\d{1,3})", text)
    if match:
        code = f"{int(match.group(1)):03d}"
        return f"Sucursal {code}", code
    return text, ""


def find_header(rows: list[tuple[Any, ...]]) -> int | None:
    for i, row in enumerate(rows[:25]):
        values = [clean(cell).lower() for cell in row[:25]]
        if "sucursal" in values and any("ticket" in value for value in values):
            return i
    return None


def column_index(headers: list[str], *needles: str) -> int | None:
    for needle in needles:
        needle = needle.lower()
        for i, header in enumerate(headers):
            if needle in header.lower():
                return i
    return None


def classify(sheet: str, description: str) -> tuple[str, str]:
    text = f"{sheet} {description}".lower()
    if any(word in text for word in ("aire", "aa ", "a.a", "split", "freon", "frio", "calor")):
        return "Aire Acondicionado", "Reparación"
    if any(word in text for word in ("luminaria", "luz", "luces", "tablero", "termica", "cable")):
        return "Problema Eléctrico", "Luminarias"
    if any(word in text for word in ("filtr", "gotera", "membrana", "techo")):
        return "Filtraciones", "Por lluvia"
    if any(word in text for word in ("pint", "durlock")):
        return "Pintura", "Interior"
    if any(word in text for word in ("material", "tubo led", "garrafa")):
        return "Materiales", "Solicitud de materiales"
    if any(word in text for word in ("persiana", "cortina", "ascensor", "escalera", "vidrio", "puerta")):
        return "Reparaciones", "General"
    return "Otro", "Otro"


def map_estado(value: str) -> str:
    state = clean(value)
    low = state.lower()
    if low in ("cerrado", "cerrar juli", "cancelado"):
        return "Cerrado"
    if low == "rechazado":
        return "Rechazado"
    if "progreso" in low or "comenz" in low and low != "no comenzado":
        return "En progreso"
    if low in ("no comenzado", "no iniciado", "abierto"):
        return "Nuevo"
    return "Pendiente" if state else "Nuevo"


def extract_provider_name(sheet: str, notes: list[str]) -> str:
    for note in notes:
        if note.lower().startswith("proveedor"):
            name = note.split(":", 1)[-1].strip()
            if "|" in name:
                name = name.split("|", 1)[0].strip()
            if name:
                return name
    return sheet.strip()


def build_tickets(input_path: Path, start_id: int, only_open: bool) -> tuple[list[dict[str, Any]], dict[str, Any]]:
    wb = load_workbook(input_path, data_only=True, read_only=True)
    tickets: list[dict[str, Any]] = []
    skipped: Counter[str] = Counter()
    by_sheet: Counter[str] = Counter()
    by_state: Counter[str] = Counter()
    next_id = start_id

    for worksheet in wb.worksheets:
        sheet = worksheet.title.strip()
        rows = list(worksheet.iter_rows(values_only=True))
        header_idx = find_header(rows)
        if header_idx is None:
            skipped["sin_header"] += 1
            continue

        headers = [clean(cell) for cell in rows[header_idx]]
        idx_suc = column_index(headers, "sucursal")
        idx_estado = column_index(headers, "estado")
        idx_prioridad = column_index(headers, "prioridad")
        idx_solicitud = column_index(headers, "solicitud")
        idx_ticket = column_index(headers, "ticket")
        idx_fecha_solicitud = column_index(headers, "fecha de solicitud")
        idx_fecha_realizacion = column_index(headers, "fecha de realización", "fecha de realizacion", "fecha de consulta")
        idx_provincia = column_index(headers, "provincia")
        idx_materiales = column_index(headers, "materiales", "rateriales")
        idx_ubicacion = column_index(headers, "ubicación de materiales", "ubicacion de materiales")

        notes: list[str] = []
        for row in rows[header_idx + 1 :]:
            first_cells = [clean(cell) for cell in row[:6] if clean(cell)]
            for value in first_cells:
                if value.lower().startswith(("proveedor", "celu", "sucursales")):
                    notes.append(value)

        provider = extract_provider_name(sheet, notes)
        for row_number, row in enumerate(rows[header_idx + 1 :], start=header_idx + 2):
            if not any(clean(cell) for cell in row[:12]):
                continue
            sucursal, sucursal_num = norm_sucursal(row[idx_suc] if idx_suc is not None else None)
            description = clean(row[idx_solicitud] if idx_solicitud is not None else "")
            invgate_id = clean(row[idx_ticket] if idx_ticket is not None else "")
            estado_original = clean(row[idx_estado] if idx_estado is not None else "")
            if not (sucursal and (description or invgate_id or estado_original)):
                skipped["fila_sin_ticket"] += 1
                continue

            estado = map_estado(estado_original)
            if only_open and estado in {"Cerrado", "Resuelto", "Rechazado"}:
                skipped["cerrados_por_only_open"] += 1
                continue

            categoria, subcategoria = classify(sheet, description)
            prioridad_raw = clean(row[idx_prioridad] if idx_prioridad is not None else "")
            try:
                prioridad = int(float(prioridad_raw)) if prioridad_raw else 4
            except ValueError:
                prioridad = 4
            prioridad = min(max(prioridad, 1), 4)

            creado = iso_date(row[idx_fecha_solicitud] if idx_fecha_solicitud is not None else "") or dt.datetime.now().isoformat()
            actualizado = iso_date(row[idx_fecha_realizacion] if idx_fecha_realizacion is not None else "") or creado
            materiales = clean(row[idx_materiales] if idx_materiales is not None else "")
            ubicacion_materiales = clean(row[idx_ubicacion] if idx_ubicacion is not None else "")

            ticket = {
                "id": next_id,
                "sucursal": sucursal,
                "sucursal_num": sucursal_num,
                "categoria": categoria,
                "subcategoria": subcategoria,
                "descripcion": description or f"Ticket InvGate {invgate_id}",
                "solicitante": "Migración InvGate",
                "prioridad": prioridad,
                "estado": estado,
                "asignado": provider if provider != "Personal Mto." else "Equipo Central",
                "fotos": [],
                "observaciones": "",
                "creado": creado,
                "actualizado": actualizado,
                "origen": "InvGate",
                "invgate_ticket_id": invgate_id,
                "invgate_estado_original": estado_original,
                "invgate_hoja": sheet,
                "invgate_fila_excel": row_number,
                "proveedor_presupuesto": provider,
                "provincia_origen": clean(row[idx_provincia] if idx_provincia is not None else ""),
                "materiales_origen": materiales,
                "ubicacion_materiales_origen": ubicacion_materiales,
                "notas": [
                    {
                        "autor": "Migración InvGate",
                        "fecha": dt.datetime.now().isoformat(),
                        "texto": f"Importado desde Excel {input_path.name}, hoja {sheet}, fila {row_number}."
                        + (f" Ticket InvGate #{invgate_id}." if invgate_id else ""),
                    }
                ],
            }
            tickets.append(ticket)
            by_sheet[sheet] += 1
            by_state[estado] += 1
            next_id += 1

    report = {
        "source": str(input_path),
        "total_tickets": len(tickets),
        "by_sheet": dict(sorted(by_sheet.items())),
        "by_state": dict(sorted(by_state.items())),
        "skipped": dict(sorted(skipped.items())),
        "only_open": only_open,
        "start_id": start_id,
    }
    return tickets, report


def load_existing(path: Path) -> list[dict[str, Any]]:
    if not path.exists():
        return []
    return json.loads(path.read_text(encoding="utf-8"))


def source_key(ticket: dict[str, Any]) -> tuple[str, str, str]:
    return (
        str(ticket.get("origen", "")),
        str(ticket.get("invgate_ticket_id", "")),
        f"{ticket.get('invgate_hoja', '')}:{ticket.get('invgate_fila_excel', '')}",
    )


def main() -> None:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("input", type=Path, help="InvGate Excel workbook")
    parser.add_argument("--output", type=Path, default=Path("data/tickets_invgate_migration_draft.json"))
    parser.add_argument("--tickets-file", type=Path, default=Path("data/tickets.json"))
    parser.add_argument("--start-id", type=int, default=100001)
    parser.add_argument("--only-open", action="store_true", help="Skip closed/rejected tickets")
    parser.add_argument("--apply", action="store_true", help="Merge into --tickets-file after writing the draft")
    args = parser.parse_args()

    tickets, report = build_tickets(args.input, args.start_id, args.only_open)
    args.output.parent.mkdir(parents=True, exist_ok=True)
    args.output.write_text(json.dumps(tickets, ensure_ascii=False, indent=2), encoding="utf-8")

    print(json.dumps(report, ensure_ascii=False, indent=2))
    print(f"draft: {args.output}")

    if not args.apply:
        return

    existing = load_existing(args.tickets_file)
    existing_keys = {source_key(ticket) for ticket in existing}
    max_id = max([int(ticket.get("id", 0)) for ticket in existing if str(ticket.get("id", "")).isdigit()] + [0])
    next_id = max(max_id + 1, args.start_id)
    to_add = []
    for ticket in tickets:
        if source_key(ticket) in existing_keys:
            continue
        ticket["id"] = next_id
        to_add.append(ticket)
        next_id += 1

    timestamp = dt.datetime.now().strftime("%Y%m%d_%H%M%S")
    backup = args.tickets_file.with_name(f"{args.tickets_file.stem}_backup_pre_invgate_{timestamp}{args.tickets_file.suffix}")
    if args.tickets_file.exists():
        shutil.copy2(args.tickets_file, backup)
    merged = existing + to_add
    args.tickets_file.write_text(json.dumps(merged, ensure_ascii=False, indent=2), encoding="utf-8")
    print(f"applied: {len(to_add)} added, {len(existing)} existing, backup: {backup}")


if __name__ == "__main__":
    main()
